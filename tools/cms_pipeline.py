#!/usr/bin/env python3
"""Valida Distrito_Go_CMS.xlsx y genera los JSON estáticos de Distrito Goo.

No se usa en el navegador ni requiere backend. Está pensado para ejecución local o CI.
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_HEADERS = {
    "Informativo": ["ID", "Actividad", "Descripción", "Link /Imagen", "Frecuencia", "Prioridad", "Categoría", "Icono", "Color", "Visible"],
    "WFM": ["Regla WFM"],
    "BT": ["Mes", "SBX", "NO. EMPLEADO", "NOMBRE COMPLETO", "CECO", "TIENDA", "REGION", "DM", "HRBP", "RD", "GB180", "ANTIGÜEDAD", "JORNADA", "ESTATUS ALTA"],
    "TBW": ["Corte", "SBX", "NOMBRE", "CeCo", "TIENDA", "PUESTO", "Región", "DM", "HRBP", "Fecha de ingreso", "Días de antigüedad", "To be Welcoming fundacional \nDía 36 al 60", "To be Welcoming Sesgo de edad\nDía 60 al 90", "To be Welcoming Discapacidad \nDía 60 al 90", "To be Welcoming Género\nDía 60 al 90", "To be Welcoming Sexualidad\ndía 90- 120", "Avance"],
    "SS": ["Mes", "SBX", "NO. EMPLEADO", "NOMBRE COMPLETO", "CECO", "TIENDA", "REGION", "DM", "HRBP", "RD", "mes de solicitud", "BT", "ESTATUS ALTA"],
    "Links": ["Categoria", "Grupo", "Icono", "Nombre", "Tipo", "URL", "WebURL", "Package", "PlayStore", "Notas", "Favorito", "Orden"],
    "Eventos": ["Fecha Inicio", "Fecha Fin", "Actividad", "Contexto / Recordatorio", "Link/Imagen", "Imagen"],
    "Actividades_Semanales": ["ID", "Actividad", "Descripción", "Día", "Hora / Corte", "Icono", "Color", "Link"],
    "Actividades_Diaria": ["ID", "Actividad", "Descripción", "Link / Imagen", "Frecuencia", "Prioridad", "Categoría", "Icono", "Color", "Visible"],
    "Duty_Roster": ["Orden", "Día", "Estaciones", "Imágenes", "Color", "Enfoque"],
    "Duty_Detail": ["Día", "Estación", "Categoría", "Orden", "Actividad", "Icono", "Crítico"],
    "Checklist_Apertura": ["Actividad", "Orden", "Concepto", "Icono"],
    "Identidad": ["Identificador", "Sección", "Campo", "Valor", "Color", "Estilo", "Visible", "Notas"],
    "Celebraciones": ["NUM_EMP", "NOMBRE", "TIENDA", "CECO", "PUESTO", "DM", "F_NAC", "F_INGRESO"],
}

CATEGORY_META = {
    "operacion": ("Operación", "⚙️", "#006241"),
    "personas": ("Personas", "👥", "#00754A"),
    "resultados": ("Resultados", "📊", "#1E6A8D"),
    "soporte": ("Soporte", "🛟", "#8A5A00"),
    "aprendizaje": ("Aprendizaje", "🎓", "#6B4EA0"),
    "otros": ("Otros", "🔗", "#5F6368"),
}


def plain(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def truthy(value: Any) -> bool:
    return str(value).strip().casefold() in {"true", "verdadero", "si", "sí", "1", "yes"}


def slug(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().casefold()
    return re.sub(r"[^a-z0-9]+", "-", text).strip("-") or "item"


def short(text: Any, limit: int = 118) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    return value if len(value) <= limit else value[: limit - 1].rstrip() + "…"


def read_sheet(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    output = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(v not in (None, "") for v in row):
            continue
        record = {headers[i]: plain(row[i]) if i < len(row) else "" for i in range(len(headers))}
        record["__row__"] = row_number
        output.append(record)
    return output


def validate_workbook(path: Path) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    errors: list[str] = []
    sheets: dict[str, list[dict[str, Any]]] = {}
    for name, required in REQUIRED_HEADERS.items():
        if name not in wb.sheetnames:
            errors.append(f"Falta la pestaña obligatoria: {name}")
            continue
        ws = wb[name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        actual = [str(v).strip() if v is not None else "" for v in first]
        missing = [h for h in required if h not in actual]
        if missing:
            errors.append(f"{name}: faltan encabezados: {', '.join(missing)}")
        sheets[name] = read_sheet(ws)
    return sheets, errors


def image_path(value: Any, root: Path, folder: str = "assets/photos") -> str:
    name = Path(str(value or "").strip()).name
    if not name:
        return ""
    base = root / folder
    matches = {p.name.casefold(): p for p in base.glob("*") if p.is_file()}
    match = matches.get(name.casefold())
    return match.relative_to(root).as_posix() if match else f"{folder}/{name}"


def resource(value: Any, root: Path) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "", ""
    if re.match(r"^https?://", text, re.I):
        return text, "link"
    return image_path(text, root), "imagen"


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build(root: Path, sheets: dict[str, list[dict[str, Any]]]) -> list[Path]:
    data = root / "data"
    changed: list[Path] = []

    def emit(name: str, value: Any):
        target = data / name
        before = target.read_text(encoding="utf-8") if target.exists() else None
        text = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
        if text != before:
            target.write_text(text, encoding="utf-8")
            changed.append(target)

    raw = {name: [{k: v for k, v in row.items() if k != "__row__"} for row in rows] for name, rows in sheets.items()}

    daily = []
    for row in raw["Actividades_Diaria"]:
        rec, typ = resource(row.pop("Link / Imagen", ""), root)
        row["ID"] = int(row["ID"])
        row["Prioridad"] = int(row["Prioridad"])
        row["Visible"] = truthy(row["Visible"])
        row.update({"Recurso": rec, "TipoRecurso": typ, "DescripcionBreve": short(row.get("Descripción"))})
        daily.append(row)

    info = []
    for row in raw["Informativo"]:
        rec, typ = resource(row.pop("Link /Imagen", ""), root)
        row["ID"] = int(row["ID"])
        row["Prioridad"] = int(row["Prioridad"])
        row["Visible"] = truthy(row["Visible"])
        row.update({"Recurso": rec, "TipoRecurso": typ, "DescripcionBreve": short(row.get("Descripción"))})
        info.append(row)

    weekly = raw["Actividades_Semanales"]
    for row in weekly:
        row["ID"] = int(row["ID"])

    events = []
    for row in raw["Eventos"]:
        mixed = str(row.pop("Link/Imagen", "") or "").strip()
        link = mixed if re.match(r"^https?://", mixed, re.I) else ""
        img_candidate = "" if link else mixed
        events.append({**row, "Link": link, "ImagenPath": image_path(img_candidate, root) if img_candidate else ""})

    duty_roster = []
    for row in raw["Duty_Roster"]:
        names = [x.strip() for x in str(row.get("Imágenes", "")).split(",") if x.strip()]
        paths = [image_path(x, root, "assets/premium/duty-roster") for x in names]
        duty_roster.append({**row, "Orden": int(row["Orden"]), "ImagenesPath": paths, "Premium": bool(paths), "Link": paths[0] if paths else ""})

    duty_detail = raw["Duty_Detail"]
    for row in duty_detail:
        row["Orden"] = int(row["Orden"])
        row["Crítico"] = truthy(row["Crítico"])

    checklist = raw["Checklist_Apertura"]
    for row in checklist:
        row["Orden"] = int(row["Orden"])

    links = raw["Links"]
    old_tools = json.loads((data / "herramientas.v10.json").read_text(encoding="utf-8")) if (data / "herramientas.v10.json").exists() else []
    old_tools_by_name = {str(t.get("nombre", "")).strip().casefold(): t for t in old_tools}
    old_categories = json.loads((data / "categorias.v10.json").read_text(encoding="utf-8")) if (data / "categorias.v10.json").exists() else []
    old_categories_by_id = {c.get("id"): c for c in old_categories}
    tools = []
    used_ids: set[str] = set()
    for row in links:
        previous = old_tools_by_name.get(str(row["Nombre"]).strip().casefold(), {})
        base_id = previous.get("id") or slug(row["Nombre"])
        item_id = base_id
        suffix = 2
        while item_id in used_ids:
            item_id = f"{base_id}-{suffix}"
            suffix += 1
        used_ids.add(item_id)
        cat_id = previous.get("categoriaId") or slug(row["Categoria"])
        if cat_id not in CATEGORY_META:
            CATEGORY_META.setdefault(cat_id, (str(row["Categoria"]), "🔗", "#5F6368"))
        keyword_source = " ".join(str(row.get(k, "")) for k in ("Categoria", "Grupo", "Nombre", "Tipo", "Notas"))
        keywords = sorted({slug(word) for word in re.findall(r"[\wÁÉÍÓÚÜÑáéíóúüñ]+", keyword_source) if len(word) > 2})
        order_value = row.get("Orden")
        order = int(order_value) if order_value not in (None, "") else int(previous.get("orden") or len(tools) + 1)
        tools.append({
            "id": item_id, "categoriaId": cat_id, "categoria": row["Categoria"], "categoriaIcono": previous.get("categoriaIcono") or old_categories_by_id.get(cat_id, {}).get("icono") or CATEGORY_META[cat_id][1],
            "grupo": row["Grupo"], "icono": row["Icono"], "nombre": row["Nombre"], "tipo": str(row["Tipo"]).casefold(),
            "url": row.get("URL", ""), "webUrl": row.get("WebURL", ""), "package": row.get("Package", ""), "playStore": row.get("PlayStore", ""),
            "notas": row.get("Notas", ""), "favorito": truthy(row.get("Favorito", "")), "orden": order, "keywords": keywords, "estado": "activo"
        })
    tools.sort(key=lambda x: (x["orden"], x["nombre"].casefold()))
    favorites = [t["id"] for t in tools if t["favorito"]]
    categories = []
    for cat_id in dict.fromkeys(t["categoriaId"] for t in tools):
        previous_category = old_categories_by_id.get(cat_id, {})
        name, icon, color = CATEGORY_META[cat_id]
        categories.append({
            "id": cat_id,
            "nombre": previous_category.get("nombre", name),
            "icono": previous_category.get("icono", icon),
            "color": previous_category.get("color", color),
            "descripcion": previous_category.get("descripcion", f"Herramientas de {name.casefold()}."),
            "contador": sum(t["categoriaId"] == cat_id for t in tools),
            "accent": previous_category.get("accent", color),
        })

    identity_rows = {r["Identificador"]: r for r in raw["Identidad"] if truthy(r["Visible"])}
    old_identity = json.loads((data / "identity.json").read_text(encoding="utf-8")) if (data / "identity.json").exists() else {}
    identity = {
        "hero": {
            **old_identity.get("hero", {}),
            "greeting": {
                "morning": identity_rows.get("hero.greeting.morning", {}).get("Valor", "Buenos días Partners."),
                "afternoon": identity_rows.get("hero.greeting.afternoon", {}).get("Valor", "Buenas tardes Partners."),
                "evening": identity_rows.get("hero.greeting.evening", {}).get("Valor", "Buenas noches Partners."),
            },
            "campaign": {
                "primary": identity_rows.get("hero.campaign.primary", {}).get("Valor", ""),
                "accent": identity_rows.get("hero.campaign.accent", {}).get("Valor", ""),
                "display": identity_rows.get("hero.campaign.display", {}).get("Valor", ""),
                "primaryColor": identity_rows.get("hero.campaign.primary", {}).get("Color", "#006241"),
                "accentColor": identity_rows.get("hero.campaign.accent", {}).get("Color", "#111111"),
                "style": identity_rows.get("hero.campaign.display", {}).get("Estilo", "corporativo-expresivo"),
                "featured": True,
            },
            "hashtags": old_identity.get("hero", {}).get("hashtags", ["#GreenApronService", "#DistritoKike"]),
        }
    }

    bt, ss, tbw = raw["BT"], raw["SS"], raw["TBW"]
    celebrations = raw["Celebraciones"]
    operational = {
        "eventos": events, "actividadesDiarias": daily, "actividadesSemanales": weekly,
        "dutyRoster": duty_roster, "dutyDetail": duty_detail, "checklistApertura": checklist,
        "altasCurso": {"bt": bt, "ss": ss, "tbw": tbw}, "wfm": raw["WFM"],
        "cmsFuente": "Distrito_Go_CMS.xlsx", "informativo": info, "celebraciones": celebrations,
        "wfmRegla": raw["WFM"][0].get("Regla WFM", "") if raw["WFM"] else "",
    }

    outputs = {
        "actividades-diarias.v10.json": daily, "actividades-semanales.v10.json": weekly,
        "informativo.v10.json": info, "eventos.v10.json": events,
        "duty-roster.v10.json": duty_roster, "duty-detail.v10.json": duty_detail,
        "checklist-apertura.v10.json": checklist, "bt.json": bt, "ss.json": ss, "tbw.json": tbw,
        "wfm.json": raw["WFM"], "links.json": links, "herramientas.v10.json": tools,
        "favoritos.v10.json": favorites, "categorias.v10.json": categories,
        "altas-curso.v10.json": {"bt": bt, "ss": ss, "tbw": tbw}, "identity.json": identity,
        "operacional.v10.json": operational,
    }
    for name, value in outputs.items():
        emit(name, value)
    return changed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("cms", type=Path)
    parser.add_argument("--project", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()
    sheets, errors = validate_workbook(args.cms)
    if errors:
        print("\n".join(f"ERROR: {e}" for e in errors))
        return 1
    print("CMS válido:", ", ".join(f"{k}={len(v)}" for k, v in sheets.items()))
    if not args.validate_only:
        changed = build(args.project.resolve(), sheets)
        print(f"JSON actualizados: {len(changed)}")
        for path in changed:
            print(path.relative_to(args.project.resolve()).as_posix())
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
